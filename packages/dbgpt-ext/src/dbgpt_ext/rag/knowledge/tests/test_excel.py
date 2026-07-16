import openpyxl
import pytest

from ..excel import ExcelKnowledge


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["name", "value"])
    ws1.append(["a", 1])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["name", "value"])
    ws2.append(["b", 2])
    wb.save(path)
    return str(path)


@pytest.fixture
def numeric_only_xlsx(tmp_path):
    path = tmp_path / "numeric_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NumericSheet"
    for r in range(6):
        ws.append([r, r + 1, r + 2])
    wb.save(path)
    return str(path)


def test_load_reads_all_sheets(multi_sheet_xlsx):
    knowledge = ExcelKnowledge(file_path=multi_sheet_xlsx)
    docs = knowledge._load()
    sheets_seen = {doc.metadata["sheet_name"] for doc in docs}
    assert sheets_seen == {"Sheet1", "Sheet2"}


def test_load_does_not_crash_without_header_row(numeric_only_xlsx):
    knowledge = ExcelKnowledge(file_path=numeric_only_xlsx)
    docs = knowledge._load()
    assert len(docs) == 6
